"""Manufacturing-spec Excel writer.

Writes the same payload the PDF writer consumes, but in a
spreadsheet-friendly shape that ERPs and supplier-portals can
ingest. Three sheets:

- **Specs** — flat key / value / unit / tolerance rows covering
  the electrical + mechanical summary.
- **BOM** — vendor PN, qty, $/unit, line total. Generated from
  ``core.cost_per_piece`` and ``wire.cost_per_meter`` when
  present.
- **Tests** — the acceptance-test plan, one row per test, with
  the same columns the PDF lays out.

Uses ``openpyxl`` (already a top-level dependency for the
existing comparison-export feature).
"""

from __future__ import annotations

from pathlib import Path

from pfc_inductor.manufacturing.spec import MfgSpec


def write_mfg_spec_xlsx(spec: MfgSpec, output_path: Path | str) -> Path:
    """Write the manufacturing spec to an XLSX workbook.

    Returns the resolved output path. The output directory is
    created if missing. Sheet column widths are tuned for a
    print-friendly A4 landscape preset.
    """
    # Lazy import — keeps the module importable in environments
    # that don't ship openpyxl (we depend on it but defensive
    # imports are cheap and they keep test scaffolding happy).
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _populate_specs_sheet(wb, spec, Alignment, Border, Font, PatternFill, Side)
    _populate_bom_sheet(wb, spec, Alignment, Border, Font, PatternFill, Side)
    _populate_tests_sheet(wb, spec, Alignment, Border, Font, PatternFill, Side)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Visual constants (mirror the PDF writer's palette)
# ---------------------------------------------------------------------------
_BAND_BG_HEX = "FFF4F4F5"
_BORDER_HEX = "FFD4D4D8"
_HEADER_FONT_HEX = "FF18181B"


def _styled_header(Alignment, Border, Font, PatternFill, Side):
    return {
        "fill": PatternFill(start_color=_BAND_BG_HEX, end_color=_BAND_BG_HEX, fill_type="solid"),
        "font": Font(name="Calibri", bold=True, size=10, color=_HEADER_FONT_HEX),
        "border": Border(
            left=Side(style="thin", color=_BORDER_HEX),
            right=Side(style="thin", color=_BORDER_HEX),
            top=Side(style="thin", color=_BORDER_HEX),
            bottom=Side(style="thin", color=_BORDER_HEX),
        ),
        "align": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


def _styled_body(Alignment, Border, Font, _PatternFill, Side):
    return {
        "font": Font(name="Calibri", size=10),
        "border": Border(
            left=Side(style="thin", color=_BORDER_HEX),
            right=Side(style="thin", color=_BORDER_HEX),
            top=Side(style="thin", color=_BORDER_HEX),
            bottom=Side(style="thin", color=_BORDER_HEX),
        ),
        "align": Alignment(horizontal="left", vertical="top", wrap_text=True),
    }


def _apply(cell, style: dict) -> None:
    if "fill" in style:
        cell.fill = style["fill"]
    cell.font = style["font"]
    cell.border = style["border"]
    cell.alignment = style["align"]


def _write_header_row(ws, row_idx: int, headers: list[str], style: dict) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=label)
        _apply(cell, style)


def _write_data_row(ws, row_idx: int, cells: list, style: dict) -> None:
    for col, value in enumerate(cells, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        _apply(cell, style)


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------
def _populate_specs_sheet(wb, spec: MfgSpec, Alignment, Border, Font, PatternFill, Side) -> None:
    """Sheet 1 — flat key/value/unit/tolerance rows."""
    ws = wb.active
    ws.title = "Specs"
    header = _styled_header(Alignment, Border, Font, PatternFill, Side)
    body = _styled_body(Alignment, Border, Font, PatternFill, Side)
    headers = ["Section", "Parameter", "Value", "Unit", "Tolerance"]
    _write_header_row(ws, 1, headers, header)

    rows: list[tuple[str, str, str, str, str]] = []
    rows.extend(
        [
            ("Project", "Project name", spec.project_name, "—", "—"),
            ("Project", "Designer", spec.designer, "—", "—"),
            ("Project", "Revision", spec.revision, "—", "—"),
            ("Project", "Date", spec.date_iso, "—", "—"),
            ("Project", "MagnaDesign ver", _magnadesign_version(), "—", "—"),
        ]
    )
    rows.extend(
        [
            ("Mechanical", "Core part", spec.core.part_number, "—", "—"),
            ("Mechanical", "Core shape", spec.core.shape, "—", "—"),
            ("Mechanical", "OD", f"{float(spec.core.OD_mm or 0):.2f}", "mm", "vendor data"),
            ("Mechanical", "ID", f"{float(spec.core.ID_mm or 0):.2f}", "mm", "vendor data"),
            ("Mechanical", "HT", f"{float(spec.core.HT_mm or 0):.2f}", "mm", "vendor data"),
            ("Mechanical", "Wire", spec.wire.id, "—", "—"),
            ("Mechanical", "Wire OD", f"{_wire_od(spec):.3f}", "mm", "vendor data"),
            ("Mechanical", "N turns", str(int(spec.result.N_turns)), "—", "exact"),
            ("Mechanical", "Layers", str(spec.winding.n_layers), "—", "exact"),
            ("Mechanical", "Bobbin used", f"{spec.winding.bobbin_used_pct:.0f}", "%", "—"),
        ]
    )
    rows.extend(
        [
            ("Electrical", "L_target", f"{spec.result.L_required_uH:.1f}", "µH", "±10 %"),
            ("Electrical", "L_actual", f"{spec.result.L_actual_uH:.1f}", "µH", "±10 %"),
            ("Electrical", "B_pk", f"{spec.result.B_pk_T * 1000:.0f}", "mT", "—"),
            ("Electrical", "P_total", f"{spec.result.losses.P_total_W:.2f}", "W", "—"),
            ("Electrical", "T_winding", f"{spec.result.T_winding_C:.1f}", "°C", "—"),
            ("Electrical", "T_rise", f"{spec.result.T_rise_C:.1f}", "°C", "—"),
        ]
    )
    rows.extend(
        [
            ("Insulation", "Class", spec.insulation.name, "—", "IEC 60085"),
            ("Insulation", "T_max", f"{spec.insulation.T_max_C:.0f}", "°C", "IEC 60085"),
            ("Insulation", "Inter-layer tape", spec.insulation.inter_layer_tape, "—", "—"),
            (
                "Insulation",
                "Tape thickness",
                f"{spec.insulation.inter_layer_tape_mm:.2f}",
                "mm",
                "—",
            ),
            ("Insulation", "Wire enamel", spec.insulation.enamel_grade, "—", "—"),
            ("Insulation", "Hi-pot voltage", f"{spec.hipot_V:.0f}", "V AC", "IEC 61558"),
            ("Insulation", "Hi-pot dwell", f"{spec.insulation.hipot_dwell_s:.0f}", "s", "—"),
        ]
    )

    for idx, row in enumerate(rows, start=2):
        _write_data_row(ws, idx, list(row), body)

    # Column widths tuned for legibility.
    widths = (16, 26, 22, 10, 22)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w


def _populate_bom_sheet(wb, spec: MfgSpec, Alignment, Border, Font, PatternFill, Side) -> None:
    """Sheet 2 — bill of materials with line totals."""
    ws = wb.create_sheet(title="BOM")
    header = _styled_header(Alignment, Border, Font, PatternFill, Side)
    body = _styled_body(Alignment, Border, Font, PatternFill, Side)
    headers = ["Item", "Vendor PN", "Description", "Qty", "Unit", "$/unit", "Line total"]
    _write_header_row(ws, 1, headers, header)

    rows: list[tuple] = []

    # Core line.
    core_cost = float(getattr(spec.core, "cost_per_piece", 0) or 0)
    rows.append(
        (
            1,
            spec.core.part_number,
            f"{spec.material.name} — {spec.core.shape}",
            1,
            "pc",
            f"{core_cost:.2f}" if core_cost else "—",
            f"{core_cost:.2f}" if core_cost else "—",
        )
    )

    # Wire line — length = N · MLT.
    wire_length_m = spec.result.N_turns * float(spec.core.MLT_mm or 0) / 1000.0
    wire_cost_per_m = float(getattr(spec.wire, "cost_per_meter", 0) or 0)
    wire_total = wire_length_m * wire_cost_per_m
    rows.append(
        (
            2,
            spec.wire.id,
            f"{spec.wire.type} wire",
            f"{wire_length_m:.2f}",
            "m",
            f"{wire_cost_per_m:.3f}" if wire_cost_per_m else "—",
            f"{wire_total:.2f}" if wire_total else "—",
        )
    )

    # Insulation tape — rough length estimate: per-layer tape =
    # one full circumference per layer.
    n_layers = max(0, spec.winding.n_layers - 1)
    tape_length_m = n_layers * float(spec.core.MLT_mm or 0) / 1000.0
    rows.append(
        (
            3,
            f"{spec.insulation.inter_layer_tape} tape {spec.insulation.inter_layer_tape_mm:.2f} mm",
            f"{spec.insulation.name} inter-layer dielectric",
            f"{tape_length_m:.2f}",
            "m",
            "—",
            "—",
        )
    )

    for idx, row in enumerate(rows, start=2):
        _write_data_row(ws, idx, list(row), body)

    widths = (6, 28, 40, 10, 8, 12, 14)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w


def _populate_tests_sheet(wb, spec: MfgSpec, Alignment, Border, Font, PatternFill, Side) -> None:
    """Sheet 3 — acceptance test plan, one row per test."""
    ws = wb.create_sheet(title="Tests")
    header = _styled_header(Alignment, Border, Font, PatternFill, Side)
    body = _styled_body(Alignment, Border, Font, PatternFill, Side)
    headers = ["#", "Test", "Condition", "Expected", "Tolerance", "Instrument"]
    _write_header_row(ws, 1, headers, header)

    for idx, test in enumerate(spec.acceptance_tests, start=1):
        _write_data_row(
            ws,
            idx + 1,
            [
                idx,
                test.name,
                test.condition,
                test.expected,
                test.tolerance,
                test.instrument,
            ],
            body,
        )

    widths = (5, 24, 30, 26, 22, 30)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w


def _wire_od(spec: MfgSpec) -> float:
    try:
        return spec.wire.outer_diameter_mm()
    except ValueError:
        return float(spec.wire.d_cu_mm or spec.wire.d_iso_mm or 0.0)


def _magnadesign_version() -> str:
    try:
        from importlib.metadata import version as _v

        return _v("magnadesign")
    except Exception:
        return "unknown"

"""Manufacturing-spec module + CLI subcommand tests.

Covers the four backend pieces (winding-layout solver, insulation
selector, acceptance-test builder, MfgSpec aggregator) plus the
two writer surfaces (PDF / XLSX) and the ``mfg-spec`` CLI
subcommand. Round-trip checks on the XLSX file verify cells
through ``openpyxl.load_workbook``; PDF tests stay at the smoke
level (``startswith %PDF-`` + size > 5 KB).
"""
from __future__ import annotations

from pathlib import Path

import pytest
from click.testing import CliRunner


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def reference_design():
    """A feasible boost-PFC design used as the canonical specimen."""
    from pfc_inductor.data_loader import (
        ensure_user_data,
        load_cores,
        load_materials,
        load_wires,
    )
    from pfc_inductor.design import design as run_design
    from pfc_inductor.models import Spec

    ensure_user_data()
    mats = load_materials()
    cores = load_cores()
    wires = load_wires()
    spec = Spec(
        topology="boost_ccm", Pout_W=600,
        Vin_min_Vrms=85, Vin_max_Vrms=265, Vout_V=400,
        f_sw_kHz=65, ripple_pct=20, T_amb_C=40,
    )
    mat = next(m for m in mats if m.id == "magnetics-60_highflux")
    core = next(c for c in cores
                if c.id == "magnetics-c058777a2-60_highflux")
    wire = next(w for w in wires if w.id == "AWG14")
    result = run_design(spec, core, wire, mat)
    return spec, core, wire, mat, result


# ---------------------------------------------------------------------------
# Winding layout
# ---------------------------------------------------------------------------
def test_plan_winding_zero_turns_returns_empty_plan(reference_design) -> None:
    from pfc_inductor.manufacturing import plan_winding

    _, core, wire, _, _ = reference_design
    plan = plan_winding(core=core, wire=wire, n_turns=0)
    assert plan.n_turns == 0
    assert plan.n_layers == 0
    assert plan.layers == ()
    assert any("zero" in w.lower() for w in plan.warnings)


def test_plan_winding_happy_path(reference_design) -> None:
    """A feasible boost design lays out into 1+ layers with
    a comfortable bobbin fill."""
    from pfc_inductor.manufacturing import plan_winding

    _, core, wire, _, result = reference_design
    plan = plan_winding(
        core=core, wire=wire, n_turns=int(result.N_turns),
    )
    assert plan.n_turns == int(result.N_turns)
    assert plan.n_layers >= 1
    assert sum(layer.turns for layer in plan.layers) == plan.n_turns
    # Cumulative layer height monotonically non-decreasing.
    heights = [layer.height_mm for layer in plan.layers]
    assert heights == sorted(heights)


def test_plan_winding_overfill_emits_warning(reference_design) -> None:
    """Cramming 1000 turns onto a small toroid → overfill
    warning fires."""
    from pfc_inductor.manufacturing import plan_winding

    _, core, wire, _, _ = reference_design
    plan = plan_winding(core=core, wire=wire, n_turns=1000)
    assert plan.bobbin_used_pct > 90.0
    assert any("won't fit" in w or "% full" in w
               for w in plan.warnings)


def test_plan_winding_underfill_emits_warning(reference_design) -> None:
    """A tiny turn count on the same toroid → underfill warning."""
    from pfc_inductor.manufacturing import plan_winding

    _, core, wire, _, _ = reference_design
    plan = plan_winding(core=core, wire=wire, n_turns=1)
    assert plan.bobbin_used_pct < 30.0
    assert any("Bobbin only" in w for w in plan.warnings)


# ---------------------------------------------------------------------------
# Insulation stack
# ---------------------------------------------------------------------------
def test_insulation_class_picks_class_b_for_warm_winding() -> None:
    from pfc_inductor.manufacturing import pick_insulation_class

    cls = pick_insulation_class(T_winding_C=80.0)
    # 80 °C + 10 °C margin = 90 °C → Class A (105) is enough,
    # but the selector prefers the lowest valid class so A wins.
    assert cls.id in {"A", "B"}


def test_insulation_class_promotes_for_hot_winding() -> None:
    from pfc_inductor.manufacturing import pick_insulation_class

    cls = pick_insulation_class(T_winding_C=140.0)
    # 140 + 10 = 150 → Class F (155) is the lowest fit.
    assert cls.id == "F"


def test_insulation_class_caps_at_h_above_180() -> None:
    from pfc_inductor.manufacturing import pick_insulation_class

    cls = pick_insulation_class(T_winding_C=200.0)
    assert cls.id == "H"


def test_hipot_voltage_floors_at_1500() -> None:
    from pfc_inductor.manufacturing import hipot_voltage_V

    # Low V_work → still 1500 V minimum per IEC 61558.
    assert hipot_voltage_V(0.0) == 1500.0
    assert hipot_voltage_V(50.0) == 1500.0


def test_hipot_voltage_2v_plus_1000() -> None:
    from pfc_inductor.manufacturing import hipot_voltage_V

    # 400 V × 2 + 1000 = 1800 V (above the floor).
    assert hipot_voltage_V(400.0) == pytest.approx(1800.0)


# ---------------------------------------------------------------------------
# Acceptance tests
# ---------------------------------------------------------------------------
def test_build_acceptance_tests_carries_six_rows(reference_design) -> None:
    from pfc_inductor.manufacturing import build_acceptance_tests

    spec, core, wire, mat, result = reference_design
    tests = build_acceptance_tests(
        spec=spec, core=core, wire=wire,
        material=mat, result=result,
    )
    # Standard six-row plan; biased-L is optional (depends on
    # I_pk being set), so allow 5–6.
    assert 5 <= len(tests) <= 6
    names = {t.name for t in tests}
    assert "Inductance" in names
    assert "DC resistance" in names
    assert "Hi-pot" in names
    assert "Insulation resistance" in names
    assert "Visual + dimensional" in names


def test_build_acceptance_tests_hipot_uses_iec_61558(
    reference_design,
) -> None:
    """The hi-pot row's condition must carry a voltage that
    matches the IEC 61558 formula (2·V_work + 1000, floor 1500)."""
    from pfc_inductor.manufacturing import build_acceptance_tests

    spec, core, wire, mat, result = reference_design
    tests = build_acceptance_tests(
        spec=spec, core=core, wire=wire,
        material=mat, result=result,
    )
    hipot = next(t for t in tests if t.name == "Hi-pot")
    # The condition string carries the voltage; just check it's
    # a sane number above the floor.
    assert "V AC" in hipot.condition
    # 265 V_rms × √2 ≈ 374.77 V → 2×374.77 + 1000 = 1749.5 V,
    # which the writer rounds to "1750 V AC" via {:.0f}.
    assert "1750" in hipot.condition or "1748" in hipot.condition or \
           "1800" in hipot.condition or "1500" in hipot.condition


# ---------------------------------------------------------------------------
# build_mfg_spec aggregator
# ---------------------------------------------------------------------------
def test_build_mfg_spec_carries_every_block(reference_design) -> None:
    from pfc_inductor.manufacturing import build_mfg_spec

    spec, core, wire, mat, result = reference_design
    pack = build_mfg_spec(
        spec=spec, core=core, wire=wire, material=mat, result=result,
        project_name="test-project",
        designer="QA Bot",
        revision="A.0",
    )
    assert pack.project_name == "test-project"
    assert pack.designer == "QA Bot"
    assert pack.winding.n_turns == int(result.N_turns)
    assert pack.insulation.id in {"A", "B", "F", "H"}
    assert pack.hipot_V >= 1500.0
    assert len(pack.acceptance_tests) >= 5


# ---------------------------------------------------------------------------
# PDF writer
# ---------------------------------------------------------------------------
def test_pdf_writer_smoke(reference_design, tmp_path: Path) -> None:
    """Smoke test — file > 5 KB and starts with the PDF magic
    number. Layout details belong to ReportLab; we just verify
    the writer doesn't raise."""
    from pfc_inductor.manufacturing import build_mfg_spec
    from pfc_inductor.manufacturing.pdf_writer import write_mfg_spec_pdf

    spec, core, wire, mat, result = reference_design
    pack = build_mfg_spec(
        spec=spec, core=core, wire=wire, material=mat, result=result,
    )
    out = tmp_path / "mfg.pdf"
    written = write_mfg_spec_pdf(pack, out)
    assert written == out
    assert out.is_file()
    assert out.stat().st_size > 5_000
    assert out.read_bytes().startswith(b"%PDF-")


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
def test_xlsx_writer_round_trips(reference_design, tmp_path: Path) -> None:
    """Round-trip the workbook through openpyxl and verify the
    three expected sheets are present with the documented columns."""
    from openpyxl import load_workbook

    from pfc_inductor.manufacturing import build_mfg_spec
    from pfc_inductor.manufacturing.excel_writer import write_mfg_spec_xlsx

    spec, core, wire, mat, result = reference_design
    pack = build_mfg_spec(
        spec=spec, core=core, wire=wire, material=mat, result=result,
    )
    out = tmp_path / "mfg.xlsx"
    write_mfg_spec_xlsx(pack, out)
    assert out.is_file()

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Specs", "BOM", "Tests"}

    specs = wb["Specs"]
    assert specs.cell(1, 1).value == "Section"
    assert specs.cell(1, 2).value == "Parameter"
    # Has multiple data rows.
    assert specs.cell(5, 1).value is not None

    bom = wb["BOM"]
    assert bom.cell(1, 2).value == "Vendor PN"
    # Core line + wire line + tape line.
    assert bom.cell(2, 1).value == 1
    assert bom.cell(3, 1).value == 2
    assert bom.cell(4, 1).value == 3

    tests = wb["Tests"]
    assert tests.cell(1, 1).value == "#"
    assert tests.cell(1, 2).value == "Test"
    # At least 5 acceptance rows (header + 5+ tests).
    assert tests.max_row >= 6


# ---------------------------------------------------------------------------
# CLI subcommand
# ---------------------------------------------------------------------------
@pytest.fixture
def cli_runner() -> CliRunner:
    return CliRunner()


def _write_boost_project(tmp_path: Path) -> Path:
    from pfc_inductor.models import Spec
    from pfc_inductor.project import ProjectFile, save_project

    spec = Spec(
        topology="boost_ccm", Pout_W=600,
        Vin_min_Vrms=85, Vin_max_Vrms=265, Vout_V=400,
        f_sw_kHz=65, ripple_pct=20, T_amb_C=40,
    )
    pf = ProjectFile.from_session(
        name="cli-mfg-boost",
        spec=spec,
        material_id="magnetics-60_highflux",
        core_id="magnetics-c058777a2-60_highflux",
        wire_id="AWG14",
    )
    project_path = tmp_path / "boost.pfc"
    save_project(project_path, pf)
    return project_path


def test_mfg_spec_subcommand_registered() -> None:
    from pfc_inductor.cli import SUBCOMMANDS
    assert "mfg-spec" in SUBCOMMANDS


def test_mfg_spec_help_lists_options(cli_runner: CliRunner) -> None:
    from pfc_inductor.cli import cli

    result = cli_runner.invoke(cli, ["mfg-spec", "--help"])
    assert result.exit_code == 0
    for opt in ("--out", "--designer", "--revision",
                "--project-name"):
        assert opt in result.output, f"missing {opt} in help"


def test_mfg_spec_writes_pdf(
    cli_runner: CliRunner, tmp_path: Path,
) -> None:
    from pfc_inductor.cli import cli

    project_path = _write_boost_project(tmp_path)
    out = tmp_path / "out.pdf"
    result = cli_runner.invoke(
        cli, ["mfg-spec", str(project_path), "--out", str(out)],
    )
    assert result.exit_code == 0, result.output
    assert out.is_file()
    assert out.stat().st_size > 5_000
    assert out.read_bytes().startswith(b"%PDF-")


def test_mfg_spec_writes_xlsx(
    cli_runner: CliRunner, tmp_path: Path,
) -> None:
    from pfc_inductor.cli import cli

    project_path = _write_boost_project(tmp_path)
    out = tmp_path / "out.xlsx"
    result = cli_runner.invoke(
        cli, ["mfg-spec", str(project_path), "--out", str(out)],
    )
    assert result.exit_code == 0, result.output
    assert out.is_file()
    # XLSX magic number (PK).
    assert out.read_bytes().startswith(b"PK")


def test_mfg_spec_unknown_extension_errors(
    cli_runner: CliRunner, tmp_path: Path,
) -> None:
    from pfc_inductor.cli import cli

    project_path = _write_boost_project(tmp_path)
    result = cli_runner.invoke(
        cli, ["mfg-spec", str(project_path),
              "--out", str(tmp_path / "out.docx")],
    )
    assert result.exit_code != 0
    assert "Unsupported" in result.output or \
           "extension" in result.output

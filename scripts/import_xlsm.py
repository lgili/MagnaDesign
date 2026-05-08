"""Convert legacy `Otimizador_Magneticos.xlsm` into our JSON databases.

Run from project root:
    .venv/bin/python scripts/import_xlsm.py

Generates:
    data/materials.json   (with fitted Steinmetz + raw datapoints + rolloff)
    data/cores.json       (~1000 parts: Magnetics, Thornton, TDK, Micrometals, Magmattec)
    data/wires.json       (full AWG table)
"""

from __future__ import annotations

import json
import math
from pathlib import Path

import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSM = ROOT / "Otimizador_Magneticos.xlsm"
DATA = ROOT / "data"


def fit_steinmetz(F_kHz, B_T, P_mWcm3) -> tuple[float, float, float]:
    """Fit P = k * f^a * B^b in log space. Returns (k, a, b) in (mW/cm^3, kHz, T).

    Clamps alpha and beta to physically reasonable ranges to avoid fit pathologies
    from narrow datasets. alpha must be >= 1.0 (Steinmetz hysteresis lower bound),
    beta typically in [1.5, 4.0] for ferrites/powder; allow up to 6.0 for HF
    materials with steep B dependence.
    """
    pts = [
        (f, b, p)
        for f, b, p in zip(F_kHz, B_T, P_mWcm3)
        if f and b and p and p > 0 and f > 0 and b > 0
    ]
    if len(pts) < 3:
        return (0.0, 1.4, 2.5)
    A = np.array([[1.0, math.log(f), math.log(b)] for f, b, _ in pts])
    y = np.array([math.log(p) for _, _, p in pts])
    sol, *_ = np.linalg.lstsq(A, y, rcond=None)
    log_k, a, b = sol
    # Clamp to physical bounds, then refit k for the clamped (a, b)
    a_clamped = float(min(max(a, 1.0), 3.0))
    b_clamped = float(min(max(b, 1.5), 6.0))
    if a_clamped != a or b_clamped != b:
        log_k = float(np.mean(y - a_clamped * A[:, 1] - b_clamped * A[:, 2]))
    return (math.exp(log_k), a_clamped, b_clamped)


def fitted_to_anchored(
    k: float, a: float, b: float, f_ref_kHz: float = 100.0, B_ref_mT: float = 100.0
) -> dict:
    """Convert P = k * f[kHz]^a * B[T]^b to anchored form Pv_ref @ (f_ref kHz, B_ref mT)."""
    B_ref_T = B_ref_mT / 1000.0
    Pv_ref = k * (f_ref_kHz**a) * (B_ref_T**b)
    return {
        "Pv_ref_mWcm3": float(Pv_ref),
        "f_ref_kHz": f_ref_kHz,
        "B_ref_mT": B_ref_mT,
        "alpha": float(a),
        "beta": float(b),
        "f_min_kHz": 1.0,
        "f_max_kHz": 500.0,
    }


# ---- Demo bulk costs (USD/kg) -------------------------------------------
# Approximate raw-material prices for visualisation. NOT production data.
# User can override per-entry via the DB editor.
def demo_cost_per_kg(vendor: str, family: str, mu_r: float | None) -> float | None:
    f = (family or "").lower()
    v = (vendor or "").lower()
    if "nano" in f:
        return 80.0
    if "ferrite" in f or "mnzn" in f:
        return 12.0
    if "highflux" in f or "high flux" in f:
        return 22.0
    if "xflux" in f or "kool" in f or "sendust" in f or "mpp" in f:
        return 18.0
    if "iron powder" in f or "powder iron" in f:
        return 8.0
    if mu_r is not None and mu_r <= 200:
        return 15.0  # generic powder fallback
    if mu_r is not None and mu_r >= 1000:
        return 12.0  # generic ferrite fallback
    return None


# Enameled magnet wire bulk price (USD/kg) including drawing + insulation.
DEMO_WIRE_COST_USD_PER_KG = 18.0


def normalize_vendor(name: str) -> str:
    """Normalize vendor names (fix typos in source spreadsheet)."""
    s = name.strip()
    return {"Thorthon": "Thornton"}.get(s, s)


def material_type(vendor: str, mu_r: float | None) -> str:
    """Heuristic material classification."""
    v = vendor.lower()
    if mu_r is not None and mu_r <= 200:
        return "powder"
    if "ferroxcube" in v or "tdk" in v or "magnetics" in v or "thornton" in v or "epcos" in v:
        return "ferrite" if mu_r and mu_r >= 800 else "powder"
    if "magmattec" in v:
        return "powder" if mu_r and mu_r <= 200 else "ferrite"
    if "micrometals" in v:
        return "powder"
    return "ferrite"


def import_materials(wb) -> dict:
    ws = wb["Material"]
    headers = [c.value for c in ws[1]]
    materials = []
    for r_idx in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r_idx]]
        vendor_raw, mat_name, mu_r, Bmax_G, Bmax_T = row[0], row[1], row[2], row[3], row[4]
        if not vendor_raw or not mat_name:
            continue
        try:
            mu_r = float(mu_r) if mu_r is not None else None
        except (TypeError, ValueError):
            mu_r = None
        if Bmax_T is None and Bmax_G is not None:
            Bmax_T = float(Bmax_G) / 10000.0  # 1 T = 10000 G
        try:
            Bmax_T = float(Bmax_T) if Bmax_T is not None else 0.0
        except (TypeError, ValueError):
            Bmax_T = 0.0
        # Fallback for materials with no Bsat in source data (e.g. Thornton ELM4)
        if Bmax_T <= 0.0:
            Bmax_T = 0.3

        # Extract loss datapoints (12 sets of F, B, P)
        F = [row[5 + i] for i in range(12)]
        B = [row[17 + i] for i in range(12)]
        P = [row[29 + i] for i in range(12)]

        F_kHz = [float(f) if f is not None else None for f in F]
        B_T = [float(b) if b is not None else None for b in B]
        P_mWcm3 = [float(p) if p is not None else None for p in P]

        valid_pts = [(f, b, p) for f, b, p in zip(F_kHz, B_T, P_mWcm3) if f and b and p and p > 0]
        if valid_pts:
            k, a, b = fit_steinmetz(*zip(*valid_pts))
            steinmetz = fitted_to_anchored(k, a, b)
            datapoints = [{"f_kHz": f, "B_T": bv, "Pv_mWcm3": p} for f, bv, p in valid_pts]
        else:
            steinmetz = {
                "Pv_ref_mWcm3": 250,
                "f_ref_kHz": 100,
                "B_ref_mT": 100,
                "alpha": 1.4,
                "beta": 2.5,
                "f_min_kHz": 1.0,
                "f_max_kHz": 500.0,
            }
            datapoints = []

        vendor = normalize_vendor(str(vendor_raw))
        family = guess_family(vendor, mat_name)
        mat_id = f"{slug(vendor)}-{slug(mat_name)}"
        type_ = material_type(vendor, mu_r)

        cost = demo_cost_per_kg(vendor, family, mu_r)
        m = {
            "id": mat_id,
            "vendor": vendor,
            "family": family,
            "name": str(mat_name),
            "type": type_,
            "mu_initial": mu_r if mu_r is not None else 1.0,
            "Bsat_25C_T": Bmax_T,
            "Bsat_100C_T": Bmax_T * 0.85,
            "rho_kg_m3": 5000.0,
            "steinmetz": steinmetz,
            "rolloff": default_rolloff(type_, mu_r, vendor, str(mat_name)),
            "loss_datapoints": datapoints,
            "cost_per_kg": cost,
            "cost_currency": "USD",
            "notes": (
                f"Imported from Otimizador_Magneticos.xlsm. "
                f"{len(datapoints)} loss datapoint(s) used to fit Steinmetz."
                f"{' Demo cost.' if cost is not None else ''}"
            ),
        }
        materials.append(m)
    return {
        "_comment": "Imported from Otimizador_Magneticos.xlsm. Steinmetz fitted from F/B/P datapoints. Rolloff is default per material type — verify against vendor datasheet for high DC bias designs.",
        "materials": materials,
    }


# Rolloff library calibrated to published 50%-permeability bias points (H_50 in Oe)
# from vendor datasheets. Values map (vendor, family_substring, mu_r) -> H_50_Oe and shape c.
# Conversion: mu_fraction = 1/(a + b*H^c) with a=0.01.
# At H_50: a + b*H_50^c = 2  (mu = 0.5), so b = 1.99 / H_50^c.

_HF_BY_MU = {  # Magnetics High Flux: mu -> H_50 Oe (datasheet 25C curves)
    14: 700,
    26: 420,
    40: 280,
    60: 150,
    125: 70,
    147: 60,
    160: 55,
}
_XF_BY_MU = {  # Magnetics XFlux
    19: 540,
    26: 380,
    40: 250,
    60: 155,
    75: 125,
    90: 105,
}
_KOOLMU_BY_MU = {  # Magnetics Kool Mu (Sendust)
    26: 280,
    60: 110,
    75: 90,
    90: 75,
    125: 50,
    147: 45,
    160: 40,
}
_MPP_BY_MU = {  # Magnetics MPP (Moly Permalloy)
    14: 600,
    26: 360,
    60: 100,
    125: 50,
    147: 45,
    160: 40,
    200: 32,
    300: 22,
}


def _curve_from_h50(H_50_Oe: float, c: float = 1.13) -> dict:

    b = 1.99 / (H_50_Oe**c)
    return {"a": 0.01, "b": float(b), "c": c, "H_units": "Oe"}


def _closest_key(d: dict, target: float) -> int:
    return min(d.keys(), key=lambda k: abs(k - target))


def lookup_rolloff(vendor: str, name: str, mu_r: float | None) -> dict | None:
    """Calibrated rolloff lookup. Returns None if not a powder/iron material."""
    if mu_r is None or mu_r <= 0:
        return None
    n = name.lower()
    v = vendor.lower()

    # Magnetics families
    if "magnetics" in v:
        if "highflux" in n or "high flux" in n:
            return _curve_from_h50(_HF_BY_MU[_closest_key(_HF_BY_MU, mu_r)], c=1.13)
        if "xflux" in n:
            return _curve_from_h50(_XF_BY_MU[_closest_key(_XF_BY_MU, mu_r)], c=1.13)
        if "koolmu" in n or "kool" in n:
            return _curve_from_h50(_KOOLMU_BY_MU[_closest_key(_KOOLMU_BY_MU, mu_r)], c=1.13)
        if "mpp" in n:
            return _curve_from_h50(_MPP_BY_MU[_closest_key(_MPP_BY_MU, mu_r)], c=1.13)
        # Magnetics ferrites (R, P, F, T, L) - no rolloff (gapped behavior dominates)
        if mu_r >= 800:
            return None

    # Magmattec (Brazilian) - similar to Magnetics powder families per their datasheet
    if "magmattec" in v:
        # 002 (mu=10) similar to HF14, 014 (mu=14) similar to HF14
        # 018 (mu=55) similar to XFlux 60, 026 (mu=75) similar to XFlux 75
        # 034 (mu=33) similar to XFlux 40, 052 (mu=75) similar to XFlux 75
        if mu_r <= 20:
            return _curve_from_h50(620, c=1.10)
        elif mu_r <= 40:
            return _curve_from_h50(290, c=1.13)
        elif mu_r <= 65:
            return _curve_from_h50(155, c=1.13)
        else:
            return _curve_from_h50(120, c=1.13)

    # Micrometals iron/SiFe powder (-MM series). Approximate from mu_r.
    if "micrometals" in v:
        # Iron powder: very gradual rolloff at high H, lower mu materials
        if mu_r <= 20:
            return _curve_from_h50(700, c=1.10)
        elif mu_r <= 40:
            return _curve_from_h50(330, c=1.13)
        elif mu_r <= 70:
            return _curve_from_h50(170, c=1.13)
        else:
            return _curve_from_h50(115, c=1.13)

    # CSC / POCO / generic Chinese
    if any(s in v for s in ("csc", "poco", "chang sung")):
        if mu_r <= 30:
            return _curve_from_h50(380, c=1.10)
        elif mu_r <= 75:
            return _curve_from_h50(140, c=1.13)
        else:
            return _curve_from_h50(95, c=1.13)

    # Thornton: TH50/TH60 are MnZn ferrites (mu>>1000), no rolloff.
    # IP6, IP12, ELM are also ferrite-class.
    if "thornton" in v or "thorthon" in v:
        if mu_r >= 800:
            return None

    # Generic powder fallback
    if mu_r <= 30:
        return _curve_from_h50(400, c=1.10)
    elif mu_r <= 100:
        return _curve_from_h50(150, c=1.13)
    else:
        return _curve_from_h50(70, c=1.20)


def default_rolloff(
    type_: str, mu_r: float | None, vendor: str = "", name: str = ""
) -> dict | None:
    """Compatible signature with prior call site, but uses the calibrated lookup."""
    if type_ != "powder":
        return None
    return lookup_rolloff(vendor, name, mu_r)


def guess_family(vendor: str, name: str) -> str:
    n = name.lower()
    if "highflux" in n or "high flux" in n:
        return "High Flux"
    if "xflux" in n:
        return "XFlux"
    if "koolmu" in n or "kool mu" in n:
        return "Kool Mu"
    if "mpp" in n:
        return "MPP"
    if name.startswith("N") and len(name) <= 4:
        return "MnZn ferrite"
    if name in ("R", "P", "F", "T", "L"):
        return "MnZn ferrite"
    if "IP" in name or name.startswith("TH") or name.startswith("ELM"):
        return "Ferrite"
    if (
        "002" in name
        or "014" in name
        or "018" in name
        or "026" in name
        or "034" in name
        or "052" in name
    ):
        return "Powder iron"
    if "MM" in name or name.startswith("-"):
        return "Iron powder mix"
    return "Material"


def slug(s: str) -> str:
    out = []
    for ch in s.strip().lower():
        if ch.isalnum():
            out.append(ch)
        elif ch in "-_":
            out.append(ch)
        elif ch == " ":
            out.append("-")
        else:
            out.append("")
    return "".join(out) or "x"


def import_cores(wb) -> dict:
    ws = wb["Core"]
    cores = []
    seen = set()
    for r_idx in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r_idx]]
        part_num, vendor_raw, mat, shape, Ae, Aw, _AeAw, AL, le, Ve, _mu_r, lgap, MLT = row[:13]
        if not part_num or not vendor_raw:
            continue
        try:
            Ae = float(Ae) if Ae is not None else 0.0
            Aw = float(Aw) if Aw is not None else 0.0
            AL = float(AL) if AL is not None else 0.0
            le = float(le) if le is not None else 0.0
            Ve = float(Ve) if Ve is not None else 0.0
            MLT = float(MLT) if MLT is not None else 0.0
            lgap = float(lgap) if lgap is not None else 0.0
        except (TypeError, ValueError):
            continue
        if Ae <= 0 or le <= 0 or AL <= 0:
            continue
        vendor = normalize_vendor(str(vendor_raw))
        material_id = f"{slug(vendor)}-{slug(str(mat))}"
        core_id = f"{slug(vendor)}-{slug(str(part_num))}-{slug(str(mat))}"
        if core_id in seen:
            core_id += f"-{r_idx}"
        seen.add(core_id)
        c = {
            "id": core_id,
            "vendor": vendor,
            "shape": str(shape) if shape else "unknown",
            "part_number": str(part_num),
            "default_material_id": material_id,
            "Ae_mm2": Ae,
            "le_mm": le,
            "Ve_mm3": Ve,
            "Wa_mm2": Aw,
            "MLT_mm": MLT,
            "AL_nH": AL,
            "lgap_mm": lgap,
            "notes": "" if lgap == 0 else f"Gapped, l_gap = {lgap} mm",
        }
        cores.append(c)
    return {
        "_comment": f"Imported {len(cores)} cores from Otimizador_Magneticos.xlsm.",
        "cores": cores,
    }


def import_wires(wb) -> dict:
    ws = wb["Wire"]
    wires = []
    for r_idx in range(3, 60):  # Headers in row 1-2, data from row 3
        row = [c.value for c in ws[r_idx]]
        awg, d_mm, A_mm2, *_ = row[:7]
        if d_mm is None or A_mm2 is None:
            continue
        try:
            d_mm = float(d_mm)
            A_mm2 = float(A_mm2)
        except (TypeError, ValueError):
            continue
        awg_str = str(awg) if awg is not None else "?"
        # Heavy-build insulation thickness ~ 0.05-0.1 mm depending on AWG
        d_iso = d_mm + max(0.05, 0.06 * d_mm)
        try:
            awg_int = int(awg)
            wid = f"AWG{awg_int}"
        except (TypeError, ValueError):
            wid = f"AWG{awg_str}"
        # Derive demo cost: copper density 8960 kg/m³, $18/kg drawn enamelled wire.
        mass_per_meter_g = A_mm2 * 1e-6 * 8960.0 * 1000.0
        cost_per_meter = (mass_per_meter_g / 1000.0) * DEMO_WIRE_COST_USD_PER_KG
        w = {
            "id": wid,
            "type": "round",
            "awg": int(awg) if isinstance(awg, (int, float)) else None,
            "d_cu_mm": d_mm,
            "d_iso_mm": d_iso,
            "A_cu_mm2": A_mm2,
            "mass_per_meter_g": round(mass_per_meter_g, 4),
            "cost_per_meter": round(cost_per_meter, 4),
            "notes": "Imported from Otimizador_Magneticos.xlsm. Demo cost @ $18/kg.",
        }
        wires.append(w)
    return {"_comment": "Full AWG table imported from Otimizador_Magneticos.xlsm.", "wires": wires}


def main() -> None:
    print(f"Reading {XLSM.name} ...")
    wb = openpyxl.load_workbook(XLSM, data_only=True, keep_vba=False)

    print("Importing materials ...")
    materials = import_materials(wb)
    (DATA / "materials.json").write_text(
        json.dumps(materials, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"  -> {len(materials['materials'])} materials")

    print("Importing cores ...")
    cores = import_cores(wb)
    (DATA / "cores.json").write_text(
        json.dumps(cores, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"  -> {len(cores['cores'])} cores")

    print("Importing wires ...")
    wires = import_wires(wb)
    (DATA / "wires.json").write_text(
        json.dumps(wires, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"  -> {len(wires['wires'])} wires")

    print("Done.")


if __name__ == "__main__":
    main()
